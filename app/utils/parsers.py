from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile
from typing import Iterable

from docx import Document
from openpyxl import load_workbook


@dataclass
class ParsedQuestion:
    text: str
    question_type: str = "single_choice"
    answers: list[tuple[str, bool]] = field(default_factory=list)
    payload: dict = field(default_factory=dict)
    points: int = 1
    image_path: str | None = None


@dataclass
class ParsedTest:
    title: str
    description: str
    questions: list[ParsedQuestion] = field(default_factory=list)
    target_score: int = 20
    questions_per_student: int = 20


class ParseError(ValueError):
    pass


SUPPORTED_EXTENSIONS = {"txt", "docx", "xlsx", "zip"}


def parse_uploaded_tests(file_path: str) -> list[ParsedTest]:
    path = Path(file_path)
    extension = path.suffix.lower().lstrip(".")

    if extension not in SUPPORTED_EXTENSIONS:
        raise ParseError("Неподдерживаемый формат файла.")

    if extension == "zip":
        tests = parse_zip_excel_bundle(file_path)
        validate_tests(tests)
        return tests

    if extension == "xlsx":
        tests = parse_excel_file(file_path)
        validate_tests(tests)
        return tests

    if extension == "docx":
        tests = parse_docx_file(file_path)
        validate_tests(tests)
        return tests

    content = path.read_text(encoding="utf-8")
    tests = parse_text_content(content, default_title=path.stem)
    validate_tests(tests)
    return tests

def parse_docx_file(file_path: str) -> list[ParsedTest]:
    document = Document(file_path)
    source_path = Path(file_path)
    entries = extract_docx_entries(document, source_path)
    return parse_docx_entries(entries, default_title=source_path.stem)


def extract_docx_entries(document: Document, source_path: Path) -> list[dict]:
    app_dir = Path(__file__).resolve().parents[1]
    image_dir = app_dir / "static" / "uploads" / "imported_questions" / source_path.stem
    image_dir.mkdir(parents=True, exist_ok=True)

    paragraph_map = {paragraph._element: paragraph for paragraph in document.paragraphs}
    entries: list[dict] = []
    image_counter = 1

    for child in document.element.body.iterchildren():
        if not child.tag.endswith("}p"):
            continue

        paragraph = paragraph_map.get(child)
        if paragraph is None:
            continue

        text = (paragraph.text or "").strip()
        if text:
            entries.append({"type": "text", "value": text})

        embeds = child.xpath('.//*[local-name()="blip"]/@*[local-name()="embed"]')
        for embed_id in embeds:
            related_part = document.part.related_parts.get(embed_id)
            if related_part is None or not getattr(related_part, "blob", None):
                continue

            ext = Path(related_part.filename or f"image_{image_counter}.png").suffix or ".png"
            filename = f"{image_counter:03d}{ext}"
            save_path = image_dir / filename
            save_path.write_bytes(related_part.blob)

            relative_static_path = f"uploads/imported_questions/{source_path.stem}/{filename}"
            entries.append({"type": "image", "value": relative_static_path})
            image_counter += 1

    return entries


def parse_docx_entries(entries: list[dict], default_title: str) -> list[ParsedTest]:
    tests: list[ParsedTest] = []
    current_test = ParsedTest(
        title=default_title,
        description=f"Импортировано из файла {default_title}.",
    )
    current_points = 1
    current_question: dict | None = None

    def flush_question() -> None:
        nonlocal current_question, current_test
        if not current_question:
            return
        question = build_block_question(current_question, current_points)
        current_test.questions.append(question)
        current_question = None

    def flush_test() -> None:
        nonlocal current_test
        flush_question()
        if current_test.questions:
            tests.append(current_test)

    for entry in entries:
        if entry["type"] == "image":
            if current_question is not None and not current_question.get("image_path"):
                current_question["image_path"] = entry["value"]
            continue

        line = str(entry["value"]).strip()
        if not line:
            continue

        lower_line = line.lower()

        if lower_line.startswith("test:") or lower_line.startswith("тест:") or line.startswith("#"):
            flush_test()
            title = line.split(":", 1)[1].strip() if ":" in line else line.lstrip("#").strip()
            current_test = ParsedTest(
                title=title or default_title,
                description=f"Импортировано из файла {default_title}.",
            )
            current_points = 1
            current_question = None
            continue

        block_match = re.match(
            r"^\s*(\d+)\s*(балл|балла|баллов|ballik|balli|ball|points?|очко|очка|очков)\b",
            lower_line,
        )
        if block_match:
            flush_question()
            current_points = int(block_match.group(1))
            continue

        if lower_line.startswith(("настройки:", "settings:", "sozlamalar:")):
            continue

        if lower_line.startswith(("вопрос:", "question:", "savol:")):
            flush_question()
            current_question = {"text": line.split(":", 1)[1].strip()}
            continue

        if lower_line.startswith(("тип:", "type:", "turi:")):
            if current_question is None:
                raise ParseError(f"Строка типа вопроса встретилась раньше самого вопроса: {line}")
            current_question["question_type"] = normalize_import_question_type(line.split(":", 1)[1].strip())
            continue

        if lower_line.startswith(("ответы:", "answers:", "javoblar:")):
            if current_question is None:
                raise ParseError(f"Строка ответов встретилась раньше самого вопроса: {line}")
            current_question["answers"] = parse_answer_parts(line.split(":", 1)[1].split("|"))
            continue

        if lower_line.startswith(("ответ:", "answer:", "javob:")):
            if current_question is None:
                raise ParseError(f"Строка ответа встретилась раньше самого вопроса: {line}")
            answer_value = line.split(":", 1)[1].strip()
            if not answer_value.startswith("*"):
                raise ParseError("У текстового вопроса правильный ответ должен начинаться с *.")
            current_question["payload"] = {"correct_text": answer_value[1:].strip()}
            continue

        # продолжение текста вопроса
        if current_question is not None:
            current_question["text"] = f"{current_question['text']} {line}".strip()

    flush_test()

    if not tests:
        raise ParseError("В файле Word не найдено ни одного корректного теста.")

    return tests


def parse_text_content(content: str, default_title: str) -> list[ParsedTest]:
    lines = [line.rstrip() for line in content.splitlines()]
    non_empty_lines = [line.strip() for line in lines if line.strip()]

    if not non_empty_lines:
        raise ParseError("В файле не найдено ни одного корректного теста.")

    if uses_block_format(non_empty_lines):
        return parse_block_text_lines(non_empty_lines, default_title)

    return parse_legacy_text_lines(non_empty_lines, default_title)


def uses_block_format(lines: list[str]) -> bool:
    prefixes = (
        "вопрос:",
        "question:",
        "savol:",
        "тип:",
        "type:",
        "turi:",
        "ответы:",
        "answers:",
        "javoblar:",
        "ответ:",
        "answer:",
        "javob:",
    )
    return any(line.lower().startswith(prefixes) for line in lines)


def parse_legacy_text_lines(lines: list[str], default_title: str) -> list[ParsedTest]:
    tests: list[ParsedTest] = []
    current_test = ParsedTest(
        title=default_title,
        description=f"Импортировано из файла {default_title}.",
    )

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        lower_line = line.lower()
        if lower_line.startswith("test:") or lower_line.startswith("тест:") or line.startswith("#"):
            title = line.split(":", 1)[1].strip() if ":" in line else line.lstrip("#").strip()
            if current_test.questions:
                tests.append(current_test)
            current_test = ParsedTest(
                title=title or default_title,
                description=f"Импортировано из файла {default_title}.",
            )
            continue

        question = parse_legacy_question_line(line)
        current_test.questions.append(question)

    if current_test.questions:
        tests.append(current_test)

    if not tests:
        raise ParseError("В файле не найдено ни одного корректного теста.")

    return tests


def parse_legacy_question_line(line: str) -> ParsedQuestion:
    if ":" in line:
        prefix, rest = line.split(":", 1)
        prefix_norm = prefix.strip().upper()
        rest = rest.strip()

        if prefix_norm == "TEXT":
            return parse_legacy_text_question(rest)
        if prefix_norm == "MATCH":
            return parse_legacy_matching_question(rest)
        if prefix_norm == "ORDER":
            return parse_legacy_ordering_question(rest)

    return parse_legacy_choice_question(line)


def parse_legacy_choice_question(line: str) -> ParsedQuestion:
    parts = [part.strip() for part in line.split("|") if part.strip()]
    if len(parts) < 3:
        raise ParseError(
            "Строка TXT/DOCX должна содержать вопрос и минимум два варианта ответа через разделитель |."
        )

    question_text = parts[0]
    answers = parse_answer_parts(parts[1:])
    correct_count = sum(1 for _, is_correct in answers if is_correct)
    question_type = "single_choice" if correct_count == 1 else "multiple_choice"

    return ParsedQuestion(
        text=question_text,
        question_type=question_type,
        answers=answers,
        payload={},
        points=1,
    )


def parse_legacy_text_question(rest: str) -> ParsedQuestion:
    parts = [part.strip() for part in rest.split("|") if part.strip()]
    if len(parts) != 2:
        raise ParseError("Формат TEXT: Вопрос | *Правильный ответ")

    question_text = parts[0]
    answer_part = parts[1]

    if not answer_part.startswith("*"):
        raise ParseError("У текстового вопроса правильный ответ должен начинаться с *.")

    correct_text = answer_part[1:].strip()
    if not correct_text:
        raise ParseError("У текстового вопроса должен быть непустой правильный ответ.")

    return ParsedQuestion(
        text=question_text,
        question_type="text",
        answers=[],
        payload={"correct_text": correct_text},
        points=1,
    )


def parse_legacy_matching_question(rest: str) -> ParsedQuestion:
    parts = [part.strip() for part in rest.split("|") if part.strip()]
    if len(parts) < 3:
        raise ParseError("Формат MATCH: Вопрос | Лево=Право | Лево=Право")

    question_text = parts[0]
    pairs = []

    for part in parts[1:]:
        if "=" not in part:
            raise ParseError("Для MATCH каждая пара должна быть в формате Лево=Право.")
        left, right = part.split("=", 1)
        left = left.strip()
        right = right.strip()
        if not left or not right:
            raise ParseError("Для MATCH левая и правая части пары не должны быть пустыми.")
        pairs.append({"left": left, "right": right})

    if len(pairs) < 2:
        raise ParseError("Для MATCH нужно минимум две пары.")

    return ParsedQuestion(
        text=question_text,
        question_type="matching",
        answers=[],
        payload={"pairs": pairs},
        points=1,
    )


def parse_legacy_ordering_question(rest: str) -> ParsedQuestion:
    parts = [part.strip() for part in rest.split("|") if part.strip()]
    if len(parts) < 3:
        raise ParseError("Формат ORDER: Вопрос | Элемент 1 | Элемент 2 | ...")

    question_text = parts[0]
    items = parts[1:]

    if len(items) < 2:
        raise ParseError("Для ORDER нужно минимум два элемента.")

    return ParsedQuestion(
        text=question_text,
        question_type="ordering",
        answers=[],
        payload={"items": items},
        points=1,
    )


def normalize_import_question_type(raw_type: str) -> str:
    value = raw_type.strip().lower()

    aliases = {
        "single_choice": "single_choice",
        "single": "single_choice",
        "multiple_choice": "multiple_choice",
        "multiple": "multiple_choice",
        "text": "text",
        "matching": "matching",
        "ordering": "ordering",

        "один правильный ответ": "single_choice",
        "один верный ответ": "single_choice",
        "один правильный": "single_choice",
        "одиночный выбор": "single_choice",

        "несколько правильных ответов": "multiple_choice",
        "несколько верных ответов": "multiple_choice",
        "множественный выбор": "multiple_choice",
        "много правильных ответов": "multiple_choice",

        "текст": "text",
        "текстовый ответ": "text",
        "текстовый": "text",

        "сопоставление": "matching",
        "сопоставить": "matching",

        "правильный порядок": "ordering",
        "порядок": "ordering",
        "упорядочивание": "ordering",

        "bir toʻgʻri javobli": "single_choice",
        "bir toʻgʻri javob": "single_choice",
        "bir togri javobli": "single_choice",
        "bir togri javob": "single_choice",

        "koʻp toʻgʻri javobli": "multiple_choice",
        "koʻp toʻgʻri javob": "multiple_choice",
        "kop toʻgʻri javobli": "multiple_choice",
        "kop toʻgʻri javob": "multiple_choice",
        "kop togri javobli": "multiple_choice",
        "kop togri javob": "multiple_choice",

        "matn": "text",
        "matnli javob": "text",

        "moslashtirish": "matching",
        "juftlash": "matching",

        "toʻgʻri tartib": "ordering",
        "togri tartib": "ordering",
        "tartiblash": "ordering",
    }

    return aliases.get(value, value)


def parse_block_text_lines(lines: list[str], default_title: str) -> list[ParsedTest]:
    tests: list[ParsedTest] = []
    current_test = ParsedTest(
        title=default_title,
        description=f"Импортировано из файла {default_title}.",
    )
    current_points = 1
    current_question: dict | None = None

    def flush_question() -> None:
        nonlocal current_question, current_test
        if not current_question:
            return

        question = build_block_question(current_question, current_points)
        current_test.questions.append(question)
        current_question = None

    def flush_test() -> None:
        nonlocal current_test
        flush_question()
        if current_test.questions:
            tests.append(current_test)

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        lower_line = line.lower()

        if lower_line.startswith("test:") or lower_line.startswith("тест:") or line.startswith("#"):
            flush_test()
            title = line.split(":", 1)[1].strip() if ":" in line else line.lstrip("#").strip()
            current_test = ParsedTest(
                title=title or default_title,
                description=f"Импортировано из файла {default_title}.",
            )
            current_points = 1
            current_question = None
            continue

        block_match = re.match(
            r"^\s*(\d+)\s*(балл|балла|баллов|ballik|balli|ball|points?|очко|очка|очков)\b",
            lower_line,
        )
        if block_match:
            flush_question()
            current_points = int(block_match.group(1))
            continue

        if lower_line.startswith(("настройки:", "settings:", "sozlamalar:")):
            continue

        if lower_line.startswith(("вопрос:", "question:", "savol:")):
            flush_question()
            current_question = {"text": line.split(":", 1)[1].strip()}
            continue

        if lower_line.startswith(("тип:", "type:", "turi:")):
            if current_question is None:
                raise ParseError(f"Строка типа вопроса встретилась раньше вопроса: {line}")
            current_question["question_type"] = normalize_import_question_type(line.split(":", 1)[1].strip())
            continue

        if lower_line.startswith(("ответы:", "answers:", "javoblar:")):
            if current_question is None:
                raise ParseError(f"Строка ответов встретилась раньше вопроса: {line}")
            current_question["answers"] = parse_answer_parts(line.split(":", 1)[1].split("|"))
            continue

        if lower_line.startswith(("ответ:", "answer:", "javob:")):
            if current_question is None:
                raise ParseError(f"Строка ответа встретилась раньше вопроса: {line}")
            answer_value = line.split(":", 1)[1].strip()
            if not answer_value.startswith("*"):
                raise ParseError("У текстового вопроса правильный ответ должен начинаться с *.")
            current_question["payload"] = {"correct_text": answer_value[1:].strip()}
            continue

        if current_question is not None:
            current_question["text"] = f"{current_question['text']} {line}".strip()

    flush_test()

    if not tests:
        raise ParseError("В файле не найдено ни одного корректного теста.")

    return tests


def build_block_question(raw_question: dict, points: int) -> ParsedQuestion:
    question_text = str(raw_question.get("text", "")).strip()
    question_type = normalize_import_question_type(str(raw_question.get("question_type", "single_choice")).strip())

    if not question_text:
        raise ParseError("У вопроса отсутствует текст.")

    if question_type in {"single_choice", "multiple_choice"}:
        answers = raw_question.get("answers", [])
        if len(answers) < 2:
            raise ParseError(f'У вопроса "{question_text}" должно быть минимум два варианта ответа.')

        correct_count = sum(1 for _, is_correct in answers if is_correct)
        if correct_count == 0:
            raise ParseError(f'У вопроса "{question_text}" нет правильного ответа.')

        normalized_type = "single_choice" if correct_count == 1 else "multiple_choice"

        return ParsedQuestion(
            text=question_text,
            question_type=normalized_type,
            answers=answers,
            payload={},
            points=points,
            image_path=raw_question.get("image_path"),
        )

    if question_type == "text":
        payload = raw_question.get("payload", {})
        correct_text = str(payload.get("correct_text", "")).strip()
        if not correct_text:
            raise ParseError(f'У текстового вопроса "{question_text}" отсутствует правильный ответ.')

        return ParsedQuestion(
            text=question_text,
            question_type="text",
            answers=[],
            payload={"correct_text": correct_text},
            points=points,
            image_path=raw_question.get("image_path"),
        )
    if question_type == "matching":
        payload = raw_question.get("payload", {})
        pairs = payload.get("pairs", [])

        if not pairs:
            answers = raw_question.get("answers", [])
            for answer_text, _ in answers:
                if "=" not in answer_text:
                    raise ParseError(f'Для matching-вопроса "{question_text}" формат должен быть Лево=Право.')
                left, right = answer_text.split("=", 1)
                left = left.strip()
                right = right.strip()
                if not left or not right:
                    raise ParseError(f'Для matching-вопроса "{question_text}" левая и правая части пары не должны быть пустыми.')
                pairs.append({"left": left, "right": right})

        if len(pairs) < 2:
            raise ParseError(f'Для matching-вопроса "{question_text}" нужно минимум две пары.')

        return ParsedQuestion(
            text=question_text,
            question_type="matching",
            answers=[],
            payload={"pairs": pairs},
            points=points,
            image_path=raw_question.get("image_path"),
        )

    if question_type == "ordering":
        payload = raw_question.get("payload", {})
        items = payload.get("items", [])

        if not items:
            answers = raw_question.get("answers", [])
            items = [answer_text for answer_text, _ in answers if answer_text]

        if len(items) < 2:
            raise ParseError(f'Для ordering-вопроса "{question_text}" нужно минимум два элемента.')

        return ParsedQuestion(
            text=question_text,
            question_type="ordering",
            answers=[],
            payload={"items": items},
            points=points,
            image_path=raw_question.get("image_path"),
        )

    raise ParseError(f"Неподдерживаемый тип вопроса: {question_type}")


def parse_answer_parts(parts: Iterable[str]) -> list[tuple[str, bool]]:
    answers: list[tuple[str, bool]] = []
    for part in parts:
        value = str(part).strip()
        if not value:
            continue
        is_correct = value.startswith("*")
        answer_text = value[1:].strip() if is_correct else value
        if not answer_text:
            raise ParseError("Текст варианта ответа не должен быть пустым.")
        answers.append((answer_text, is_correct))
    return answers


def parse_excel_row(values: list[str], sheet_title: str, points: int = 1) -> ParsedQuestion:
    if not values:
        raise ParseError(f'Пустая строка в листе "{sheet_title}".')

    if len(values) >= 4 and values[1].lower() in {"single_choice", "multiple_choice", "text", "matching", "ordering"}:
        question_text = values[0]
        question_type = normalize_import_question_type(values[1])

        if question_type in {"single_choice", "multiple_choice"}:
            answers = parse_answer_parts(values[2:])
            if len(answers) < 2:
                raise ParseError(f'У вопроса "{question_text}" недостаточно вариантов ответа.')
            return ParsedQuestion(
                text=question_text,
                question_type=question_type,
                answers=answers,
                payload={},
                points=points,
            )

        if question_type == "text":
            correct_text = values[2].strip() if len(values) > 2 else ""
            if not correct_text.startswith("*"):
                raise ParseError(f'У текстового вопроса "{question_text}" правильный ответ должен начинаться с *.')
            return ParsedQuestion(
                text=question_text,
                question_type="text",
                answers=[],
                payload={"correct_text": correct_text[1:].strip()},
                points=points,
            )

        if question_type == "matching":
            pairs = []
            for raw_pair in values[2:]:
                if "=" not in raw_pair:
                    raise ParseError(f'Для matching-вопроса "{question_text}" каждая пара должна быть в формате Лево=Право.')
                left, right = raw_pair.split("=", 1)
                left = left.strip()
                right = right.strip()
                if not left or not right:
                    raise ParseError(f'Для matching-вопроса "{question_text}" левая и правая части пары не должны быть пустыми.')
                pairs.append({"left": left, "right": right})
            if len(pairs) < 2:
                raise ParseError(f'Для matching-вопроса "{question_text}" нужно минимум две пары.')
            return ParsedQuestion(
                text=question_text,
                question_type="matching",
                answers=[],
                payload={"pairs": pairs},
                points=points,
            )

        if question_type == "ordering":
            items = [value for value in values[2:] if value]
            if len(items) < 2:
                raise ParseError(f'Для ordering-вопроса "{question_text}" нужно минимум два элемента.')
            return ParsedQuestion(
                text=question_text,
                question_type="ordering",
                answers=[],
                payload={"items": items},
                points=points,
            )

    question_text = values[0]
    answers = parse_answer_parts(values[1:])
    if len(answers) < 2:
        raise ParseError(
            f'В листе "{sheet_title}" вопрос "{question_text}" должен содержать минимум два варианта ответа.'
        )

    correct_count = sum(1 for _, is_correct in answers if is_correct)
    question_type = "single_choice" if correct_count == 1 else "multiple_choice"

    return ParsedQuestion(
        text=question_text,
        question_type=question_type,
        answers=answers,
        payload={},
        points=points,
    )


def parse_excel_file(file_path: str) -> list[ParsedTest]:
    workbook = load_workbook(file_path, data_only=True)
    tests: list[ParsedTest] = []

    try:
        for sheet in workbook.worksheets:
            parsed_test = ParsedTest(
                title=sheet.title.strip() or "Импортированный тест",
                description=f'Импортировано из Excel-листа {sheet.title.strip() or "без названия"}.',
                target_score=20,
                questions_per_student=20,
            )

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            header_map = {header: idx for idx, header in enumerate(headers) if header}

            new_format = "question" in header_map and "type" in header_map

            current_points = 1

            for row in rows[1:] if new_format else rows:
                values = ["" if cell is None else str(cell).strip() for cell in row]

                if not any(values):
                    continue

                first_value = values[0].strip().lower() if values else ""

                if not new_format:
                    if first_value in {"type", "тип", "question", "вопрос"}:
                        continue
                    if first_value == "settings":
                        continue
                    parsed_test.questions.append(parse_excel_row([v for v in values if v], sheet.title, points=current_points))
                    continue

                block_value = get_excel_value(values, header_map, "block_points")
                if block_value:
                    if str(block_value).strip().isdigit():
                        current_points = int(str(block_value).strip())

                question_text = get_excel_value(values, header_map, "question")
                if not question_text:
                    continue

                image_name = get_excel_value(values, header_map, "image_name")
                question_type = normalize_import_question_type(get_excel_value(values, header_map, "type", "single_choice"))
                correct_raw = get_excel_value(values, header_map, "correct", "")

                parsed_question = build_excel_variant_question(
                    question_text=question_text,
                    question_type=question_type,
                    values=values,
                    header_map=header_map,
                    current_points=current_points,
                    image_name=image_name,
                    correct_raw=correct_raw,
                    source_path=Path(file_path),
                )
                parsed_test.questions.append(parsed_question)

            if parsed_test.questions:
                tests.append(parsed_test)

        if not tests:
            raise ParseError("В Excel-файле не найдено ни одного корректного теста.")

        return tests
    finally:
        workbook.close()


def get_excel_value(row: list[str], header_map: dict[str, int], key: str, default: str = "") -> str:
    index = header_map.get(key)
    if index is None:
        return default
    if index >= len(row):
        return default
    return row[index].strip()


def build_excel_variant_question(
    question_text: str,
    question_type: str,
    values: list[str],
    header_map: dict[str, int],
    current_points: int,
    image_name: str,
    correct_raw: str,
    source_path: Path,
) -> ParsedQuestion:
    image_path = None
    if image_name:
        image_path = resolve_excel_image_path(source_path, image_name)

    if question_type in {"single_choice", "multiple_choice"}:
        option_headers = sorted(
            [header for header in header_map.keys() if header.startswith("answer")],
            key=lambda item: int(re.sub(r"\D", "", item) or "0"),
        )
        answers_raw = [get_excel_value(values, header_map, header) for header in option_headers]
        answers_raw = [value for value in answers_raw if value]

        if len(answers_raw) < 2:
            raise ParseError(f'У вопроса "{question_text}" недостаточно вариантов ответа.')

        correct_indexes = set()
        for part in str(correct_raw).split(","):
            part = part.strip()
            if part.isdigit():
                correct_indexes.add(int(part))

        if not correct_indexes:
            raise ParseError(f'У вопроса "{question_text}" не указаны правильные ответы в колонке correct.')

        answers: list[tuple[str, bool]] = []
        for idx, answer in enumerate(answers_raw, start=1):
            answers.append((answer, idx in correct_indexes))

        normalized_type = "single_choice" if len(correct_indexes) == 1 else "multiple_choice"
        return ParsedQuestion(
            text=question_text,
            question_type=normalized_type,
            answers=answers,
            payload={},
            points=current_points,
            image_path=image_path,
        )

    if question_type == "text":
        correct_text = get_excel_value(values, header_map, "answer1") or correct_raw
        if not correct_text:
            raise ParseError(f'У текстового вопроса "{question_text}" отсутствует правильный ответ.')
        return ParsedQuestion(
            text=question_text,
            question_type="text",
            answers=[],
            payload={"correct_text": correct_text},
            points=current_points,
            image_path=image_path,
        )

    if question_type == "matching":
        option_headers = sorted(
            [header for header in header_map.keys() if header.startswith("answer")],
            key=lambda item: int(re.sub(r"\D", "", item) or "0"),
        )
        pairs = []
        for header in option_headers:
            raw_pair = get_excel_value(values, header_map, header)
            if not raw_pair:
                continue
            if "=" not in raw_pair:
                raise ParseError(f'Для matching-вопроса "{question_text}" каждая пара должна быть в формате Лево=Право.')
            left, right = raw_pair.split("=", 1)
            left = left.strip()
            right = right.strip()
            if not left or not right:
                raise ParseError(f'Для matching-вопроса "{question_text}" левая и правая части пары не должны быть пустыми.')
            pairs.append({"left": left, "right": right})

        if len(pairs) < 2:
            raise ParseError(f'Для matching-вопроса "{question_text}" нужно минимум две пары.')

        return ParsedQuestion(
            text=question_text,
            question_type="matching",
            answers=[],
            payload={"pairs": pairs},
            points=current_points,
            image_path=image_path,
        )

    if question_type == "ordering":
        option_headers = sorted(
            [header for header in header_map.keys() if header.startswith("answer")],
            key=lambda item: int(re.sub(r"\D", "", item) or "0"),
        )
        items = [get_excel_value(values, header_map, header) for header in option_headers]
        items = [item for item in items if item]

        if len(items) < 2:
            raise ParseError(f'Для ordering-вопроса "{question_text}" нужно минимум два элемента.')

        return ParsedQuestion(
            text=question_text,
            question_type="ordering",
            answers=[],
            payload={"items": items},
            points=current_points,
            image_path=image_path,
        )

    raise ParseError(f'Неподдерживаемый тип вопроса: {question_type}')


def resolve_excel_image_path(source_path: Path, image_name: str) -> str:
    app_dir = Path(__file__).resolve().parents[1]
    target_dir = app_dir / "static" / "uploads" / "imported_questions" / source_path.stem
    target_dir.mkdir(parents=True, exist_ok=True)

    possible_locations = [
        source_path.parent / image_name,
        source_path.parent / "images" / image_name,
    ]

    source_image = None
    for location in possible_locations:
        if location.exists() and location.is_file():
            source_image = location
            break

    if source_image is None:
        raise ParseError(f'Не найдена картинка "{image_name}" рядом с Excel-файлом.')

    target_path = target_dir / source_image.name
    if not target_path.exists():
        target_path.write_bytes(source_image.read_bytes())

    return f"uploads/imported_questions/{source_path.stem}/{target_path.name}"


def validate_tests(tests: list[ParsedTest]) -> None:
    for parsed_test in tests:
        if not parsed_test.questions:
            raise ParseError(f'Тест "{parsed_test.title}" не содержит вопросов.')

        for question in parsed_test.questions:
            if not question.text:
                raise ParseError(f'В тесте "{parsed_test.title}" найден вопрос без текста.')

def parse_group_excel(file_path: str) -> tuple[str, list[dict]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ParseError("Excel-файл группы пуст.")

        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        expected = ["group_name", "full_name", "username", "password"]
        if header[:4] != expected:
            raise ParseError("Первая строка должна быть: group_name, full_name, username, password")

        students = []
        group_name = None

        for row in rows[1:]:
            values = [str(cell).strip() if cell is not None else "" for cell in row[:4]]
            if not any(values):
                continue

            row_group, full_name, username, password = values

            if not row_group or not full_name or not username or not password:
                raise ParseError("У каждой строки должны быть заполнены group_name, full_name, username, password")

            if group_name is None:
                group_name = row_group
            elif group_name != row_group:
                raise ParseError("В одном файле должна быть только одна группа")

            students.append(
                {
                    "full_name": full_name,
                    "username": username,
                    "password": password,
                }
            )

        if not group_name:
            raise ParseError("Не удалось определить название группы.")
        if not students:
            raise ParseError("В файле группы не найдено ни одного студента.")

        return group_name, students

    finally:
        workbook.close()

def parse_zip_excel_bundle(file_path: str) -> list[ParsedTest]:
    source_zip = Path(file_path)

    with TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)

        with ZipFile(source_zip, "r") as archive:
            archive.extractall(tmp_path)

        excel_files = list(tmp_path.rglob("*.xlsx"))
        if not excel_files:
            raise ParseError("В ZIP-архиве не найден Excel-файл .xlsx.")

        if len(excel_files) > 1:
            raise ParseError("В ZIP-архиве должно быть только одно Excel-файл .xlsx.")

        excel_path = excel_files[0]
        return parse_excel_file(str(excel_path))